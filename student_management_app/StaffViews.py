import csv
import io
import json
from datetime import date, timedelta

from openpyxl import load_workbook
from django.conf import settings
from django.contrib import messages
from django.core import serializers
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.forms import model_to_dict
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone

from student_management_app.forms import StaffAddStudentForm, StaffEditStudentForm
from student_management_app.models import Subjects, SemesterModel, Students, Attendance, AttendanceReport, \
    LeaveReportStaff, Staffs, FeedBackStaffs, FeedBackStudent, CustomUser, Courses, NotificationStaffs, StudentResult, OnlineClassRoom
from student_management_app.notification_utils import (
    create_student_notifications,
    send_student_notification_emails,
)
from student_management_app.services.live_class_service import (
    create_or_get_active_room,
    end_room,
    issue_realtime_token,
    mark_participant_joined,
    serialize_room_state,
)


IMPORT_HEADERS = [
    "first_name",
    "last_name",
    "username",
    "email",
    "notification_email",
    "password",
    "address",
    "gender",
    "department",
    "semester_start_date",
    "semester_end_date",
]


def _logged_in_staff(request):
    return Staffs.objects.get(admin=request.user.id)


def _normalize_email(value):
    return (value or "").strip().lower()


def _credentials_error(username, email, exclude_user_id=None):
    username = (username or "").strip()
    email = _normalize_email(email)

    if not username or not email:
        return "Username and email are required"

    username_qs = CustomUser.objects.filter(username__iexact=username)
    email_qs = CustomUser.objects.filter(email__iexact=email)
    if exclude_user_id:
        username_qs = username_qs.exclude(id=exclude_user_id)
        email_qs = email_qs.exclude(id=exclude_user_id)

    if username_qs.exists():
        return "Username already exists"
    if email_qs.exists():
        return "Email already exists"
    return None


def _subject_for_staff_or_none(subject_id, request_user_id):
    return Subjects.objects.filter(id=subject_id, staff_id=request_user_id).first()


def _assigned_student_or_none(student_admin_id, staff_obj):
    return Students.objects.filter(
        admin=student_admin_id,
        assigned_staff=staff_obj,
    ).first()


def _student_for_staff_or_none(student_admin_id, staff_obj):
    return Students.objects.filter(
        admin=student_admin_id,
        assigned_staff=staff_obj,
    ).select_related("admin", "course_id", "semester_id", "assigned_staff__admin").first()


def _student_queryset_for_staff_list():
    return (
        Students.objects.select_related("admin", "course_id", "semester_id", "assigned_staff__admin")
        .order_by("admin__first_name", "admin__last_name", "admin__username", "id")
    )


def _resolve_semester_for_date(target_date):
    return SemesterModel.object.filter(
        semester_start_date__lte=target_date,
        semester_end_date__gte=target_date,
    ).order_by("id").first()


def _resolve_semester_from_students(subject, staff_obj):
    semester_ids = list(
        Students.objects.filter(
            course_id=subject.course_id,
            assigned_staff=staff_obj,
        ).values_list("semester_id", flat=True).distinct()
    )
    if len(semester_ids) == 1:
        return SemesterModel.object.filter(id=semester_ids[0]).first()
    return None


def _resolve_fallback_semester(attendance_date_obj, subject_model, staff_obj, selected_student_admin_ids=None):
    semester_model = _resolve_semester_for_date(attendance_date_obj)
    if semester_model:
        return semester_model

    students_qs = Students.objects.filter(
        assigned_staff=staff_obj,
        course_id=subject_model.course_id,
    )
    if selected_student_admin_ids:
        students_qs = students_qs.filter(admin_id__in=selected_student_admin_ids)

    semester_id = students_qs.values_list("semester_id", flat=True).order_by("semester_id").first()
    if semester_id:
        return SemesterModel.object.filter(id=semester_id).first()

    semester_model = SemesterModel.object.order_by("id").first()
    if semester_model:
        return semester_model

    year = attendance_date_obj.year
    return SemesterModel.object.create(
        semester_start_date=f"{year}-01-01",
        semester_end_date=f"{year}-12-31",
    )


def _course_for_name(value):
    normalized_value = " ".join((value or "").strip().lower().split())
    if not normalized_value:
        return None
    for course in Courses.objects.all():
        if " ".join(course.course_name.strip().lower().split()) == normalized_value:
            return course
    return None


def _semester_for_dates(start_value, end_value):
    try:
        start_date = date.fromisoformat((start_value or "").strip())
        end_date = date.fromisoformat((end_value or "").strip())
    except ValueError:
        try:
            start_date = date.fromisoformat(
                "-".join(reversed((start_value or "").strip().split("-")))
            )
            end_date = date.fromisoformat(
                "-".join(reversed((end_value or "").strip().split("-")))
            )
        except ValueError:
            return None

    return SemesterModel.object.filter(
        semester_start_date=start_date,
        semester_end_date=end_date,
    ).first()


def _gender_value(value):
    gender = (value or "").strip().lower()
    if gender == "male":
        return "Male"
    if gender == "female":
        return "Female"
    return None


def _create_student_user(*, first_name, last_name, username, email, notification_email, password, address, course, semester, sex):
    user = CustomUser.objects.create_user(
        username=username,
        password=password,
        email=email,
        notification_email=notification_email,
        last_name=last_name,
        first_name=first_name,
        user_type=3,
    )
    user.students.address = address
    user.students.course_id = course
    user.students.semester_id = semester
    user.students.gender = sex
    user.students.assigned_staff = None
    user.students.save()
    return user


def _extract_import_rows(uploaded_file):
    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".csv"):
        uploaded_file.seek(0)
        wrapper = io.TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
        try:
            reader = csv.DictReader(wrapper)
            headers = [header.strip().lower() for header in (reader.fieldnames or [])]
            rows = list(reader)
        finally:
            wrapper.detach()
        return headers, rows

    if filename.endswith(".xlsx"):
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet = workbook.active
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(header or "").strip().lower() for header in (raw_headers or [])]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row is None or not any(value not in (None, "") for value in row):
                continue
            rows.append(
                {
                    headers[index]: "" if value is None else str(value).strip()
                    for index, value in enumerate(row[: len(headers)])
                }
            )
        workbook.close()
        return headers, rows

    raise ValueError("Only CSV and XLSX files are supported")

def _daily_attendance_payload_for_staff(staff_user_id, selected_date):
    staff_obj = Staffs.objects.get(admin=staff_user_id)
    students = Students.objects.filter(assigned_staff=staff_obj).select_related("admin").order_by("id")
    reports = AttendanceReport.objects.filter(
        attendance_id__subject_id__staff_id=staff_user_id,
        attendance_id__attendance_date=selected_date,
        student_id__in=students,
    )
    status_map = {}
    for report in reports.select_related("student_id"):
        sid = report.student_id_id
        # If any subject report is absent for the day, treat the student as absent.
        if sid not in status_map:
            status_map[sid] = bool(report.status)
        elif not report.status:
            status_map[sid] = False

    present_students = []
    absent_students = []
    for student in students:
        display_name = (
            student.admin.get_full_name().strip()
            or student.admin.username
            or student.profile_id
        )
        if student.id not in status_map:
            continue
        is_present = status_map[student.id]
        if is_present:
            present_students.append(display_name)
        else:
            absent_students.append(display_name)

    present_students.sort()
    absent_students.sort()
    marked_total = len(present_students) + len(absent_students)

    return {
        "selected_date": selected_date,
        "previous_date": selected_date - timedelta(days=1),
        "next_date": selected_date + timedelta(days=1),
        "present_count": len(present_students),
        "total_students": marked_total,
        "present_students": present_students,
        "absent_students": absent_students,
    }


def staff_home(request):
    #For Fetch All Student Under Staff
    staff_obj = _logged_in_staff(request)
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    students_qs=Students.objects.filter(assigned_staff=staff_obj)
    students_count=students_qs.count()

    #Fetch All Attendance Count
    attendance_count=Attendance.objects.filter(subject_id__in=subjects).count()

    #Fetch All Approve Leave
    leave_count=LeaveReportStaff.objects.filter(staff_id=staff_obj.id,leave_status=1).count()
    subject_count=subjects.count()

    #Fetch Attendance Data by Subject
    subject_list=[]
    attendance_list=[]
    for subject in subjects:
        attendance_count1=Attendance.objects.filter(subject_id=subject.id).count()
        subject_list.append(subject.subject_name)
        attendance_list.append(attendance_count1)

    students_attendance=students_qs
    student_list=[]
    student_list_attendance_present=[]
    student_list_attendance_absent=[]
    for student in students_attendance:
        attendance_present_count=AttendanceReport.objects.filter(status=True,student_id=student.id).count()
        attendance_absent_count=AttendanceReport.objects.filter(status=False,student_id=student.id).count()
        student_list.append(student.admin.username)
        student_list_attendance_present.append(attendance_present_count)
        student_list_attendance_absent.append(attendance_absent_count)

    return render(request,"staff_template/staff_home_template.html",{"students_count":students_count,"attendance_count":attendance_count,"leave_count":leave_count,"subject_count":subject_count,"subject_list":subject_list,"attendance_list":attendance_list,"student_list":student_list,"present_list":student_list_attendance_present,"absent_list":student_list_attendance_absent})

def staff_manage_student(request):
    staff_obj = _logged_in_staff(request)
    students = _student_queryset_for_staff_list()
    return render(
        request,
        "staff_template/staff_manage_student.html",
        {"students": students, "staff_obj": staff_obj},
    )


def staff_students_under_me(request):
    staff_obj = _logged_in_staff(request)
    students = (
        Students.objects.filter(assigned_staff=staff_obj)
        .select_related("admin", "course_id", "semester_id")
        .order_by("admin__first_name", "admin__last_name", "admin__username", "id")
    )
    return render(
        request,
        "staff_template/staff_students_under_me.html",
        {"students": students},
    )


@require_POST
def staff_send_student_notification(request):
    staff_obj = _logged_in_staff(request)
    title = (request.POST.get("title") or "").strip()
    message = (request.POST.get("message") or "").strip()
    selected_student_ids = request.POST.getlist("student_ids")

    if not title:
        messages.error(request, "Notification title is required")
        return HttpResponseRedirect(reverse("staff_students_under_me"))

    if not message:
        messages.error(request, "Notification body is required")
        return HttpResponseRedirect(reverse("staff_students_under_me"))

    if not selected_student_ids:
        messages.error(request, "Please select at least one student")
        return HttpResponseRedirect(reverse("staff_students_under_me"))

    students = list(
        Students.objects.filter(
            assigned_staff=staff_obj,
            id__in=selected_student_ids,
        ).select_related("admin")
    )

    if not students:
        messages.error(request, "No assigned students matched your selection")
        return HttpResponseRedirect(reverse("staff_students_under_me"))

    sender_name = request.user.get_full_name().strip() or request.user.username or "Staff"
    create_student_notifications(
        students,
        sender_name=sender_name,
        title=title,
        message=message,
    )
    send_student_notification_emails(
        students,
        sender_name=sender_name,
        title=title,
        message=message,
    )
    messages.success(
        request,
        f"Notification sent to {len(students)} assigned student(s).",
    )
    return HttpResponseRedirect(reverse("staff_students_under_me"))


def staff_add_student(request):
    form = StaffAddStudentForm()
    return render(request, "staff_template/staff_add_student_template.html", {"form": form})


def staff_add_student_save(request):
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")

    form = StaffAddStudentForm(request.POST)
    if not form.is_valid():
        return render(request, "staff_template/staff_add_student_template.html", {"form": form})

    first_name = form.cleaned_data["first_name"]
    last_name = form.cleaned_data["last_name"]
    username = form.cleaned_data["username"]
    email = _normalize_email(form.cleaned_data["email"])
    notification_email = _normalize_email(form.cleaned_data["notification_email"]) or email
    password = form.cleaned_data["password"]
    address = form.cleaned_data["address"]
    semester_id = form.cleaned_data["semester_id"]
    course_id = form.cleaned_data["course"]
    sex = form.cleaned_data["sex"]

    if not password:
        messages.error(request, "Password is required")
        return HttpResponseRedirect(reverse("staff_add_student"))

    error = _credentials_error(username, email)
    if error:
        messages.error(request, error)
        return HttpResponseRedirect(reverse("staff_add_student"))

    try:
        with transaction.atomic():
            course_obj = Courses.objects.get(id=course_id)
            semester = SemesterModel.object.get(id=semester_id)
            _create_student_user(
                first_name=first_name,
                last_name=last_name,
                username=username,
                email=email,
                notification_email=notification_email,
                password=password,
                address=address,
                course=course_obj,
                semester=semester,
                sex=sex,
            )
        messages.success(request, "Successfully Added Student")
    except (Courses.DoesNotExist, SemesterModel.DoesNotExist):
        messages.error(request, "Invalid department or semester")
    except Exception:
        messages.error(request, "Failed to Add Student")
    return HttpResponseRedirect(reverse("staff_add_student"))


def staff_edit_student(request, student_id):
    staff_obj = _logged_in_staff(request)
    student = _student_for_staff_or_none(student_id, staff_obj)
    if student is None:
        messages.error(request, "You can edit only students assigned to you")
        return HttpResponseRedirect(reverse("staff_manage_student"))

    request.session["staff_student_id"] = student_id
    form = StaffEditStudentForm()
    form.fields["email"].initial = student.admin.email
    form.fields["notification_email"].initial = student.admin.notification_email
    form.fields["first_name"].initial = student.admin.first_name
    form.fields["last_name"].initial = student.admin.last_name
    form.fields["username"].initial = student.admin.username
    form.fields["address"].initial = student.address
    form.fields["course"].initial = student.course_id.id
    form.fields["sex"].initial = student.gender
    form.fields["semester_id"].initial = student.semester_id.id
    return render(
        request,
        "staff_template/staff_edit_student_template.html",
        {"form": form, "id": student_id, "username": student.admin.username},
    )


def staff_edit_student_save(request):
    if request.method != "POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")

    student_id = request.session.get("staff_student_id")
    if student_id is None:
        return HttpResponseRedirect(reverse("staff_manage_student"))

    staff_obj = _logged_in_staff(request)
    student = _student_for_staff_or_none(student_id, staff_obj)
    if student is None:
        messages.error(request, "You can edit only students assigned to you")
        return HttpResponseRedirect(reverse("staff_manage_student"))

    form = StaffEditStudentForm(request.POST)
    if not form.is_valid():
        return render(
            request,
            "staff_template/staff_edit_student_template.html",
            {"form": form, "id": student_id, "username": student.admin.username},
        )

    first_name = form.cleaned_data["first_name"]
    last_name = form.cleaned_data["last_name"]
    username = form.cleaned_data["username"]
    email = _normalize_email(form.cleaned_data["email"])
    notification_email = _normalize_email(form.cleaned_data["notification_email"]) or email
    address = form.cleaned_data["address"]
    semester_id = form.cleaned_data["semester_id"]
    course_id = form.cleaned_data["course"]
    sex = form.cleaned_data["sex"]

    error = _credentials_error(username, email, exclude_user_id=student_id)
    if error:
        messages.error(request, error)
        return HttpResponseRedirect(reverse("staff_edit_student", kwargs={"student_id": student_id}))

    try:
        with transaction.atomic():
            user = CustomUser.objects.get(id=student_id)
            user.first_name = first_name
            user.last_name = last_name
            user.username = username
            user.email = email
            user.notification_email = notification_email
            user.save()

            student = Students.objects.get(admin=student_id, assigned_staff=staff_obj)
            student.address = address
            student.semester_id = SemesterModel.object.get(id=semester_id)
            student.gender = sex
            student.course_id = Courses.objects.get(id=course_id)
            student.save()
        del request.session["staff_student_id"]
        messages.success(request, "Successfully Edited Student")
        return HttpResponseRedirect(reverse("staff_edit_student", kwargs={"student_id": student_id}))
    except (Courses.DoesNotExist, SemesterModel.DoesNotExist):
        messages.error(request, "Invalid department or semester")
    except Exception:
        messages.error(request, "Failed to Edit Student")
    return HttpResponseRedirect(reverse("staff_edit_student", kwargs={"student_id": student_id}))


def staff_delete_student(request, student_id):
    staff_obj = _logged_in_staff(request)
    student = _student_for_staff_or_none(student_id, staff_obj)
    if student is None:
        messages.error(request, "You can delete only students assigned to you")
        return HttpResponseRedirect(reverse("staff_manage_student"))

    try:
        CustomUser.objects.get(id=student_id).delete()
        messages.success(request, "Successfully Deleted Student")
    except Exception:
        messages.error(request, "Failed to Delete Student")
    return HttpResponseRedirect(reverse("staff_manage_student"))


def staff_import_students_save(request):
    if request.method != "POST":
        return HttpResponseRedirect(reverse("staff_add_student"))

    upload = request.FILES.get("student_file")
    if upload is None:
        messages.error(request, "Please choose a CSV or XLSX file to upload")
        return HttpResponseRedirect(reverse("staff_add_student"))

    try:
        headers, rows = _extract_import_rows(upload)
    except ValueError as exc:
        messages.error(request, str(exc))
        return HttpResponseRedirect(reverse("staff_add_student"))
    except Exception:
        messages.error(request, "Failed to read the uploaded file")
        return HttpResponseRedirect(reverse("staff_add_student"))

    if headers != IMPORT_HEADERS:
        messages.error(request, "Invalid file format. Please use the provided student import template.")
        return HttpResponseRedirect(reverse("staff_add_student"))

    created_count = 0
    skipped_reasons = []
    for row_index, row in enumerate(rows, start=2):
        first_name = (row.get("first_name") or "").strip()
        last_name = (row.get("last_name") or "").strip()
        username = (row.get("username") or "").strip()
        email = _normalize_email(row.get("email"))
        notification_email = _normalize_email(row.get("notification_email")) or email
        password = (row.get("password") or "").strip()
        address = (row.get("address") or "").strip()
        department_name = (row.get("department") or "").strip()
        sex = _gender_value(row.get("gender"))
        semester = _semester_for_dates(row.get("semester_start_date"), row.get("semester_end_date"))
        course = _course_for_name(department_name)

        if not all([first_name, last_name, username, email, password, address, department_name, sex]):
            skipped_reasons.append(f"Row {row_index}: missing required values")
            continue
        if course is None:
            skipped_reasons.append(f"Row {row_index}: unknown department")
            continue
        if semester is None:
            skipped_reasons.append(f"Row {row_index}: unknown semester")
            continue

        error = _credentials_error(username, email)
        if error:
            skipped_reasons.append(f"Row {row_index}: {error}")
            continue

        try:
            with transaction.atomic():
                _create_student_user(
                    first_name=first_name,
                    last_name=last_name,
                    username=username,
                    email=email,
                    notification_email=notification_email,
                    password=password,
                    address=address,
                    course=course,
                    semester=semester,
                    sex=sex,
                )
            created_count += 1
        except Exception:
            skipped_reasons.append(f"Row {row_index}: failed to create student")

    skipped_count = len(skipped_reasons)
    if created_count:
        messages.success(request, f"Imported {created_count} students successfully")
    if skipped_count:
        messages.warning(request, f"Skipped {skipped_count} rows during import")
        for reason in skipped_reasons[:10]:
            messages.warning(request, reason)
        if skipped_count > 10:
            messages.warning(request, f"{skipped_count - 10} more rows were skipped")
    if not created_count and not skipped_count:
        messages.info(request, "The file did not contain any student rows")
    return HttpResponseRedirect(reverse("staff_add_student"))


def staff_download_student_template(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="student_import_template.csv"'
    writer = csv.writer(response)
    writer.writerow(IMPORT_HEADERS)
    writer.writerow(
        [
            "Asha",
            "Patel",
            "asha.patel",
            "asha.patel@example.com",
            "asha.notifications@example.com",
            "pass12345",
            "Ahmedabad",
            "Female",
            "BCA",
            "01-01-2026",
            "30-06-2026",
        ]
    )
    return response

def staff_daily_attendance_graph(request):
    selected_date_raw = (request.GET.get("date") or "").strip()

    try:
        selected_date = date.fromisoformat(selected_date_raw) if selected_date_raw else timezone.localdate()
    except ValueError:
        selected_date = timezone.localdate()

    payload = _daily_attendance_payload_for_staff(request.user.id, selected_date)
    context = {
        "selected_date": payload["selected_date"],
        "previous_date": payload["previous_date"],
        "next_date": payload["next_date"],
        "present_count": payload["present_count"],
        "total_students": payload["total_students"],
        "present_students_json": json.dumps(payload["present_students"]),
        "absent_students_json": json.dumps(payload["absent_students"]),
    }
    return render(request, "staff_template/staff_daily_attendance_graph.html", context)

def staff_daily_attendance_graph_data(request):
    selected_date_raw = (request.GET.get("date") or "").strip()
    try:
        selected_date = date.fromisoformat(selected_date_raw) if selected_date_raw else timezone.localdate()
    except ValueError:
        return JsonResponse({"ok": False, "error": "Invalid date format"}, status=400)

    payload = _daily_attendance_payload_for_staff(request.user.id, selected_date)
    return JsonResponse(
        {
            "ok": True,
            "selected_date": payload["selected_date"].isoformat(),
            "selected_date_display": payload["selected_date"].strftime("%d-%m-%Y"),
            "previous_date": payload["previous_date"].isoformat(),
            "next_date": payload["next_date"].isoformat(),
            "present_count": payload["present_count"],
            "total_students": payload["total_students"],
            "present_students": payload["present_students"],
            "absent_students": payload["absent_students"],
        }
    )

def staff_take_attendance(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    return render(
        request,
        "staff_template/staff_take_attendance.html",
        {
            "subjects": subjects,
            "today_date": timezone.localdate().isoformat(),
        },
    )

@csrf_exempt
def get_students(request):
    subject_id=request.POST.get("subject")
    semester_id=request.POST.get("semester_id")
    attendance_date_raw = (request.POST.get("attendance_date") or "").strip()

    subject=_subject_for_staff_or_none(subject_id, request.user.id)
    if not subject:
        return JsonResponse({"error": "Invalid subject for this staff"}, status=403)

    semester_model = None
    staff_obj = _logged_in_staff(request)
    if attendance_date_raw:
        try:
            date.fromisoformat(attendance_date_raw)
        except ValueError:
            return JsonResponse({"error": "Invalid attendance date"}, status=400)
        students=Students.objects.filter(
            course_id=subject.course_id,
            assigned_staff=staff_obj,
        )
    elif semester_id:
        semester_model = SemesterModel.object.filter(id=semester_id).first()
        if not semester_model:
            semester_model = _resolve_semester_from_students(subject, staff_obj)
        if not semester_model:
            return JsonResponse(json.dumps([]), content_type="application/json", safe=False)
        students=Students.objects.filter(
            course_id=subject.course_id,
            semester_id=semester_model,
            assigned_staff=staff_obj,
        )
    else:
        students=Students.objects.filter(
            course_id=subject.course_id,
            assigned_staff=staff_obj,
        )
    list_data=[]

    for student in students:
        data_small={"id":student.admin.id,"name":student.admin.first_name+" "+student.admin.last_name}
        list_data.append(data_small)
    return JsonResponse(json.dumps(list_data),content_type="application/json",safe=False)

@csrf_exempt
def save_attendance_data(request):
    student_ids=request.POST.get("student_ids")
    subject_id=request.POST.get("subject_id")
    attendance_date=request.POST.get("attendance_date")
    subject_model=_subject_for_staff_or_none(subject_id, request.user.id)
    try:
        attendance_date_obj = date.fromisoformat(attendance_date)
    except (TypeError, ValueError):
        return HttpResponse("ERR")

    staff_obj = _logged_in_staff(request)
    json_sstudent=json.loads(student_ids)
    selected_student_admin_ids = [int(stud["id"]) for stud in json_sstudent]
    semester_model = _resolve_fallback_semester(
        attendance_date_obj,
        subject_model,
        staff_obj,
        selected_student_admin_ids=selected_student_admin_ids,
    ) if subject_model else None
    if not subject_model or not semester_model:
        return HttpResponse("ERR")

    try:
        with transaction.atomic():
            attendance, _ = Attendance.objects.get_or_create(
                subject_id=subject_model,
                attendance_date=attendance_date_obj,
                defaults={"semester_id": semester_model},
            )
            if attendance.semester_id_id != semester_model.id:
                attendance.semester_id = semester_model
                attendance.save(update_fields=["semester_id"])

            for stud in json_sstudent:
                 student=_assigned_student_or_none(stud['id'], staff_obj)
                 if student is None:
                     raise ValueError("Unauthorized student in attendance payload")
                 if student.course_id_id != subject_model.course_id_id:
                     raise ValueError("Student-course mismatch for this subject")
                 status_value = bool(int(stud['status']))
                 attendance_report, created = AttendanceReport.objects.get_or_create(
                     student_id=student,
                     attendance_id=attendance,
                     defaults={"status": status_value},
                 )
                 if not created:
                     attendance_report.status = status_value
                     attendance_report.save(update_fields=["status"])
        return HttpResponse("OK")
    except Exception:
        return HttpResponse("ERR")

def staff_update_attendance(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    return render(request,"staff_template/staff_update_attendance.html",{"subjects":subjects})

@csrf_exempt
def get_attendance_dates(request):
    subject=request.POST.get("subject")
    subject_obj=_subject_for_staff_or_none(subject, request.user.id)
    if not subject_obj:
        return JsonResponse(json.dumps([]),safe=False)

    attendance=Attendance.objects.filter(subject_id=subject_obj).order_by("-attendance_date", "-id")
    attendance_obj=[]
    for attendance_single in attendance:
        data={"id":attendance_single.id,"attendance_date":str(attendance_single.attendance_date),"semester_id":attendance_single.semester_id.id}
        attendance_obj.append(data)

    return JsonResponse(json.dumps(attendance_obj),safe=False)

@csrf_exempt
def get_attendance_student(request):
    attendance_date=request.POST.get("attendance_date")
    attendance=Attendance.objects.filter(id=attendance_date).first()
    if not attendance or attendance.subject_id.staff_id_id != request.user.id:
        return JsonResponse(json.dumps([]),content_type="application/json",safe=False)

    attendance_data=AttendanceReport.objects.filter(attendance_id=attendance)
    staff_obj = _logged_in_staff(request)
    list_data=[]

    for student in attendance_data:
        if student.student_id.assigned_staff_id != staff_obj.id:
            continue
        data_small={"id":student.student_id.admin.id,"name":student.student_id.admin.first_name+" "+student.student_id.admin.last_name,"status":student.status}
        list_data.append(data_small)
    return JsonResponse(json.dumps(list_data),content_type="application/json",safe=False)

@csrf_exempt
def save_updateattendance_data(request):
    student_ids=request.POST.get("student_ids")
    attendance_date=request.POST.get("attendance_date")
    attendance=Attendance.objects.filter(id=attendance_date).first()
    if not attendance or attendance.subject_id.staff_id_id != request.user.id:
        return HttpResponse("ERR")

    json_sstudent=json.loads(student_ids)
    staff_obj = _logged_in_staff(request)


    try:
        for stud in json_sstudent:
             student=_assigned_student_or_none(stud['id'], staff_obj)
             if student is None:
                 raise ValueError("Unauthorized student in attendance payload")
             attendance_report=AttendanceReport.objects.get(student_id=student,attendance_id=attendance)
             attendance_report.status=stud['status']
             attendance_report.save()
        return HttpResponse("OK")
    except:
        return HttpResponse("ERR")

def staff_apply_leave(request):
    staff_obj = Staffs.objects.get(admin=request.user.id)
    leave_data=LeaveReportStaff.objects.filter(staff_id=staff_obj)
    return render(request,"staff_template/staff_apply_leave.html",{"leave_data":leave_data})

def staff_apply_leave_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("staff_apply_leave"))
    else:
        leave_date=request.POST.get("leave_date")
        leave_msg=request.POST.get("leave_msg")

        staff_obj=Staffs.objects.get(admin=request.user.id)
        try:
            leave_report=LeaveReportStaff(staff_id=staff_obj,leave_date=leave_date,leave_message=leave_msg,leave_status=0)
            leave_report.save()
            messages.success(request, "Successfully Applied for Leave")
            return HttpResponseRedirect(reverse("staff_apply_leave"))
        except:
            messages.error(request, "Failed To Apply for Leave")
            return HttpResponseRedirect(reverse("staff_apply_leave"))


def staff_feedback(request):
    staff_obj = Staffs.objects.get(admin=request.user.id)
    feedback_data = FeedBackStaffs.objects.filter(staff_id=staff_obj).order_by("-created_at")
    student_feedback_data = (
        FeedBackStudent.objects.filter(staff_id=staff_obj)
        .select_related("student_id__admin", "staff_id__admin")
        .order_by("-created_at")
    )
    return render(
        request,
        "staff_template/staff_feedback.html",
        {
            "feedback_data": feedback_data,
            "student_feedback_data": student_feedback_data,
        },
    )

def staff_feedback_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("staff_feedback_save"))
    else:
        feedback_msg=request.POST.get("feedback_msg")

        staff_obj=Staffs.objects.get(admin=request.user.id)
        try:
            feedback=FeedBackStaffs(staff_id=staff_obj,feedback=feedback_msg,feedback_reply="")
            feedback.save()
            messages.success(request, "Successfully Sent Feedback")
            return HttpResponseRedirect(reverse("staff_feedback"))
        except:
            messages.error(request, "Failed To Send Feedback")
            return HttpResponseRedirect(reverse("staff_feedback"))


@require_POST
def staff_student_feedback_reply(request):
    staff_obj = Staffs.objects.get(admin=request.user.id)
    feedback_id = request.POST.get("feedback_id")
    reply_message = (request.POST.get("reply_message") or "").strip()

    if not feedback_id or not reply_message:
        messages.error(request, "Reply message cannot be empty")
        return HttpResponseRedirect(reverse("staff_feedback"))

    feedback = (
        FeedBackStudent.objects.filter(id=feedback_id, staff_id=staff_obj)
        .select_related("student_id__admin")
        .first()
    )
    if feedback is None:
        messages.error(request, "Student feedback not found")
        return HttpResponseRedirect(reverse("staff_feedback"))

    feedback.feedback_reply = reply_message
    feedback.save(update_fields=["feedback_reply"])
    messages.success(
        request,
        f"Reply saved for {feedback.student_id.admin.get_full_name() or feedback.student_id.admin.username}",
    )
    return HttpResponseRedirect(reverse("staff_feedback"))


@require_POST
def staff_student_feedback_forward(request):
    staff_obj = Staffs.objects.get(admin=request.user.id)
    feedback_id = request.POST.get("feedback_id")

    feedback = (
        FeedBackStudent.objects.filter(id=feedback_id, staff_id=staff_obj)
        .select_related("student_id__admin")
        .first()
    )
    if feedback is None:
        messages.error(request, "Student feedback not found")
        return HttpResponseRedirect(reverse("staff_feedback"))

    if feedback.forwarded_to_hod:
        messages.info(request, "This feedback has already been forwarded to HOD")
        return HttpResponseRedirect(reverse("staff_feedback"))

    feedback.forwarded_to_hod = True
    feedback.forwarded_at = timezone.now()
    feedback.save(update_fields=["forwarded_to_hod", "forwarded_at"])
    messages.success(
        request,
        f"Forwarded {feedback.student_id.admin.get_full_name() or feedback.student_id.admin.username}'s feedback to HOD",
    )
    return HttpResponseRedirect(reverse("staff_feedback"))

def staff_profile(request):
    user=CustomUser.objects.get(id=request.user.id)
    staff=Staffs.objects.get(admin=user)
    return render(request,"staff_template/staff_profile.html",{"user":user,"staff":staff})

def staff_profile_save(request):
    if request.method!="POST":
        return HttpResponseRedirect(reverse("staff_profile"))
    else:
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        address=request.POST.get("address")
        notification_email = _normalize_email(request.POST.get("notification_email"))
        password=request.POST.get("password")
        try:
            customuser=CustomUser.objects.get(id=request.user.id)
            customuser.first_name=first_name
            customuser.last_name=last_name
            customuser.notification_email = notification_email or customuser.email
            if password!=None and password!="":
                customuser.set_password(password)
            customuser.save()

            staff=Staffs.objects.get(admin=customuser.id)
            staff.address=address
            staff.save()
            messages.success(request, "Successfully Updated Profile")
            return HttpResponseRedirect(reverse("staff_profile"))
        except:
            messages.error(request, "Failed to Update Profile")
            return HttpResponseRedirect(reverse("staff_profile"))

@csrf_exempt
def staff_fcmtoken_save(request):
    token=request.POST.get("token")
    try:
        staff=Staffs.objects.get(admin=request.user.id)
        staff.fcm_token=token
        staff.save()
        return HttpResponse("True")
    except:
        return HttpResponse("False")

def staff_all_notification(request):
    staff=Staffs.objects.get(admin=request.user.id)
    notifications=NotificationStaffs.objects.filter(staff_id=staff.id).order_by("-created_at")
    return render(request,"staff_template/all_notification.html",{"notifications":notifications})

def staff_add_result(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    semesters=SemesterModel.object.all()
    return render(request,"staff_template/staff_add_result.html",{"subjects":subjects,"semesters":semesters})

def save_student_result(request):
    if request.method!='POST':
        return HttpResponseRedirect('staff_add_result')
    student_admin_id=request.POST.get('student_list')
    assignment_marks=request.POST.get('assignment_marks')
    exam_marks=request.POST.get('exam_marks')
    subject_id=request.POST.get('subject')


    staff_obj = _logged_in_staff(request)
    student_obj=_assigned_student_or_none(student_admin_id, staff_obj)
    subject_obj=_subject_for_staff_or_none(subject_id, request.user.id)
    if student_obj is None or subject_obj is None:
        messages.error(request, "Invalid student or subject assignment")
        return HttpResponseRedirect(reverse("staff_add_result"))

    try:
        check_exist=StudentResult.objects.filter(subject_id=subject_obj,student_id=student_obj).exists()
        if check_exist:
            result=StudentResult.objects.get(subject_id=subject_obj,student_id=student_obj)
            result.subject_assignment_marks=assignment_marks
            result.subject_exam_marks=exam_marks
            result.save()
            messages.success(request, "Successfully Updated Result")
            return HttpResponseRedirect(reverse("staff_add_result"))
        else:
            result=StudentResult(student_id=student_obj,subject_id=subject_obj,subject_exam_marks=exam_marks,subject_assignment_marks=assignment_marks)
            result.save()
            messages.success(request, "Successfully Added Result")
            return HttpResponseRedirect(reverse("staff_add_result"))
    except:
        messages.error(request, "Failed to Add Result")
        return HttpResponseRedirect(reverse("staff_add_result"))

@csrf_exempt
def fetch_result_student(request):
    subject_id=request.POST.get('subject_id')
    student_id=request.POST.get('student_id')
    staff_obj = _logged_in_staff(request)
    student_obj=_assigned_student_or_none(student_id, staff_obj)
    subject_obj=_subject_for_staff_or_none(subject_id, request.user.id)
    if student_obj is None or subject_obj is None:
        return HttpResponse("False")
    result=StudentResult.objects.filter(student_id=student_obj.id,subject_id=subject_obj.id).exists()
    if result:
        result=StudentResult.objects.get(student_id=student_obj.id,subject_id=subject_obj.id)
        result_data={"exam_marks":result.subject_exam_marks,"assign_marks":result.subject_assignment_marks}
        return HttpResponse(json.dumps(result_data))
    else:
        return HttpResponse("False")

def start_live_classroom(request):
    subjects=Subjects.objects.filter(staff_id=request.user.id)
    semesters=SemesterModel.object.all()
    return render(request,"staff_template/start_live_classroom.html",{"subjects":subjects,"semesters":semesters})

def start_live_classroom_process(request):
    semester_id=request.POST.get("semester_id")
    subject=request.POST.get("subject")

    room = create_or_get_active_room(request.user, subject, semester_id)
    mark_participant_joined(request.user, room, "STAFF", is_publisher=True)
    return render(
        request,
        "staff_template/live_class_room_start.html",
        {
            "username": request.user.username,
            "password": room.room_pwd,
            "roomid": room.room_name,
            "subject": room.subject.subject_name,
            "semester": room.semester,
            "room_pk": room.id,
            "socket_url": settings.LIVE_SIGNALING_URL,
        },
    )


@require_POST
def start_live_classroom_api(request):
    if str(request.user.user_type) != "2":
        return JsonResponse({"ok": False, "error": "Only staff can start class"}, status=403)

    subject = request.POST.get("subject")
    semester_id = request.POST.get("semester_id")
    if not subject or not semester_id:
        return JsonResponse({"ok": False, "error": "subject and semester_id are required"}, status=400)

    try:
        room = create_or_get_active_room(request.user, subject, semester_id)
        mark_participant_joined(request.user, room, "STAFF", is_publisher=True)
    except ObjectDoesNotExist:
        return JsonResponse({"ok": False, "error": "Invalid subject/semester"}, status=404)

    token = issue_realtime_token(request.user, room, role="STAFF")
    return JsonResponse(
        {
            "ok": True,
            "room": serialize_room_state(room),
            "room_name": room.room_name,
            "room_password": room.room_pwd,
            "token": token,
            "socket_url": settings.LIVE_SIGNALING_URL,
        }
    )


@require_POST
def end_live_classroom_api(request, room_id):
    if str(request.user.user_type) != "2":
        return JsonResponse({"ok": False, "error": "Only staff can end class"}, status=403)

    try:
        room = OnlineClassRoom.objects.get(id=room_id)
        if room.started_by.admin_id != request.user.id:
            return JsonResponse({"ok": False, "error": "Only class owner can end this room"}, status=403)
        end_room(request.user, room)
    except ObjectDoesNotExist:
        return JsonResponse({"ok": False, "error": "Room not found"}, status=404)

    return JsonResponse({"ok": True, "room": serialize_room_state(room)})


@require_POST
@csrf_exempt
def save_live_class_snapshot_api(request, room_id):
    if str(request.user.user_type) != "2":
        return JsonResponse({"ok": False, "error": "Only staff can save snapshot"}, status=403)

    try:
        room = OnlineClassRoom.objects.get(id=room_id)
        if room.started_by.admin_id != request.user.id:
            return JsonResponse({"ok": False, "error": "Only class owner can save snapshot"}, status=403)
    except ObjectDoesNotExist:
        return JsonResponse({"ok": False, "error": "Room not found"}, status=404)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except (TypeError, ValueError, json.JSONDecodeError):
        return JsonResponse({"ok": False, "error": "Invalid JSON body"}, status=400)

    snapshot = payload.get("snapshot")
    if snapshot is None:
        return JsonResponse({"ok": False, "error": "snapshot is required"}, status=400)

    room.last_board_snapshot = json.dumps(snapshot)
    room.save(update_fields=["last_board_snapshot"])
    return JsonResponse({"ok": True})


def returnHtmlWidget(request):
    return render(request,"widget.html")
